from __future__ import annotations

import io
import json
import mimetypes
import os
import re
from email.parser import BytesParser
from email.policy import default
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber


ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
DEFAULT_PORT = 8765
MAX_UPLOAD_BYTES = 25 * 1024 * 1024


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True


def _num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("¥", "").replace("₩", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _display_formula_or_value(cell: Any) -> Any:
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        return value
    return value


def _is_additional_cost_row(name: str, option: str, mark: str, shipped_qty: float) -> bool:
    text = f"{name} {option}".lower()
    keywords = [
        "바코드",
        "barcode",
        "원산지",
        "자재비",
        "자재비용",
        "인쇄",
        "택배비",
        "라벨",
        "스티커",
        "포장비",
        "부자재",
        "발급비",
    ]
    return not mark and shipped_qty == 0 and any(keyword in text for keyword in keywords)


def _allocate_additional_costs(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    items: list[dict[str, Any]] = []
    additional_costs: list[dict[str, Any]] = []
    pending_products: list[dict[str, Any]] = []
    pending_costs: list[dict[str, Any]] = []

    def flush_group() -> None:
        nonlocal pending_products, pending_costs
        if not pending_products:
            additional_costs.extend(pending_costs)
            pending_costs = []
            return

        product_total = sum(row["quantity"] * row["unitCny"] + row["inlandCny"] for row in pending_products)
        quantity_total = sum(row["quantity"] for row in pending_products)
        add_total = sum(cost["amountCny"] for cost in pending_costs)

        for product in pending_products:
            base = product["quantity"] * product["unitCny"] + product["inlandCny"]
            if product_total:
                ratio = base / product_total
            elif quantity_total:
                ratio = product["quantity"] / quantity_total
            else:
                ratio = 1 / len(pending_products)
            product["extraCny"] = round(add_total * ratio, 6)
            product["extraCostNote"] = ", ".join(cost["label"] for cost in pending_costs)
            items.append(product)

        additional_costs.extend(pending_costs)
        pending_products = []
        pending_costs = []

    for row in rows:
        if row["kind"] == "product":
            pending_products.append(row["item"])
        elif row["kind"] == "additional":
            pending_costs.append(row["cost"])
        elif row["kind"] == "new_group":
            flush_group()

    flush_group()
    return items, additional_costs


def parse_shipping_list(file_bytes: bytes, filename: str) -> dict[str, Any]:
    wb_formula = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    wb_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    detail_name = "detail-A" if "detail-A" in wb_formula.sheetnames else wb_formula.sheetnames[0]
    detail_formula = wb_formula[detail_name]
    detail_values = wb_values[detail_name]

    parsed_rows: list[dict[str, Any]] = []
    current_order = ""
    current_mark = ""
    current_name = ""
    current_url = ""

    for row_idx in range(6, detail_formula.max_row + 1):
        first = _text(detail_formula.cell(row_idx, 1).value)
        if first.upper() == "TOTAL":
            break

        explicit_order_no = _text(detail_formula.cell(row_idx, 1).value)
        if explicit_order_no and current_order:
            parsed_rows.append({"kind": "new_group"})

        order_no = explicit_order_no or current_order
        mark = _text(detail_formula.cell(row_idx, 2).value) or current_mark
        url = _text(detail_formula.cell(row_idx, 3).value) or current_url
        name = _text(detail_formula.cell(row_idx, 5).value) or current_name
        option = _text(detail_formula.cell(row_idx, 6).value)
        quantity = _num(detail_values.cell(row_idx, 7).value)
        unit_cny = _num(detail_values.cell(row_idx, 8).value)
        inland_cny = _num(detail_values.cell(row_idx, 14).value)
        shipped_qty = _num(detail_values.cell(row_idx, 18).value) or quantity

        if order_no:
            current_order = order_no
        if mark:
            current_mark = mark
        if url:
            current_url = url
        if name:
            current_name = name

        has_cost_signal = any([quantity, unit_cny, inland_cny])
        if not has_cost_signal:
            continue

        raw_name = _text(detail_formula.cell(row_idx, 5).value)
        raw_mark = _text(detail_formula.cell(row_idx, 2).value)
        amount_cny = quantity * unit_cny + inland_cny
        if _is_additional_cost_row(raw_name or name, option, raw_mark, _num(detail_values.cell(row_idx, 18).value)):
            parsed_rows.append(
                {
                    "kind": "additional",
                    "cost": {
                        "row": row_idx,
                        "orderNo": order_no,
                        "label": option or raw_name or f"부대비용 {row_idx}",
                        "amountCny": amount_cny,
                    },
                }
            )
            continue

        parsed_rows.append(
            {
                "kind": "product",
                "item": {
                    "id": f"item-{row_idx}",
                    "row": row_idx,
                    "orderNo": order_no,
                    "mark": mark,
                    "url": url,
                    "name": name or option or f"Row {row_idx}",
                    "option": option,
                    "quantity": quantity,
                    "unitCny": unit_cny,
                    "inlandCny": inland_cny,
                    "extraCny": 0,
                    "extraCostNote": "",
                    "shippedQty": shipped_qty,
                    "memo": _text(detail_formula.cell(row_idx, 11).value),
                },
            }
        )

    items, additional_costs = _allocate_additional_costs(parsed_rows)

    shipping: list[dict[str, Any]] = []
    if "출고리스트" in wb_formula.sheetnames:
        ship_formula = wb_formula["출고리스트"]
        ship_values = wb_values["출고리스트"]
        last_title = ""
        last_en_title = ""
        for row_idx in range(10, ship_formula.max_row + 1):
            first = _text(ship_formula.cell(row_idx, 1).value)
            if "GRAND TOTAL" in first.upper():
                break

            title = _text(ship_formula.cell(row_idx, 3).value) or last_title
            en_title = _text(ship_formula.cell(row_idx, 4).value) or last_en_title
            if title:
                last_title = title
            if en_title:
                last_en_title = en_title

            no = _display_formula_or_value(ship_formula.cell(row_idx, 1))
            if no is None and not title:
                continue

            shipping.append(
                {
                    "row": row_idx,
                    "no": no,
                    "mark": _text(ship_formula.cell(row_idx, 2).value),
                    "title": title,
                    "englishTitle": en_title,
                    "material": _text(ship_formula.cell(row_idx, 5).value),
                    "spec": _text(ship_formula.cell(row_idx, 6).value),
                    "cartons": _num(ship_values.cell(row_idx, 7).value),
                    "pieces": _num(ship_values.cell(row_idx, 8).value),
                    "kgs": _num(ship_values.cell(row_idx, 9).value),
                    "cbm": _num(ship_values.cell(row_idx, 10).value),
                    "declarationUnit": _num(ship_values.cell(row_idx, 11).value),
                    "declarationTotal": _num(ship_values.cell(row_idx, 12).value),
                }
            )

    meta = {
        "filename": filename,
        "sheets": wb_formula.sheetnames,
        "detailSheet": detail_name,
        "shippingSheet": "출고리스트" if "출고리스트" in wb_formula.sheetnames else "",
        "additionalCostCount": len(additional_costs),
        "additionalCostCny": sum(cost["amountCny"] for cost in additional_costs),
    }

    return {"meta": meta, "items": items, "shipping": shipping, "additionalCosts": additional_costs}


def parse_invoice_pdf(file_bytes: bytes, filename: str) -> dict[str, Any]:
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        text_parts: list[str] = []
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
            for table in page.extract_tables() or []:
                for row in table:
                    cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if cells:
                        text_parts.append(" ".join(cells))
        text = "\n".join(text_parts)

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    costs = {"tax": [], "logistics": [], "other": []}
    meta: dict[str, Any] = {"filename": filename}
    section = ""

    label_map = {
        "OCEAN FREIGHT CHARGE": "해상운임",
        "TERMINAL HANDLING": "터미널 핸들링",
        "DOCUMENT FEE": "서류 발급비",
        "C.F.S CHARGE": "CFS 비용",
        "HANDLING CHARGE": "핸들링 비용",
        "OTHER CHARGE": "기타 물류비",
        "STORAGE CHARGE": "창고료",
        "CERTIFICATE OF ORIGIN": "원산지 증명서",
        "CUSTOMS CLEARANCE": "통관수수료",
        "INLAND CHARGE": "국내 운송비",
        "TRUCKING": "국내 운송비",
        "DELIVERY": "배송비",
        "WAREHOUSE": "창고료",
        "WHARFAGE": "부두사용료",
        "DO CHARGE": "D/O 비용",
        "D/O CHARGE": "D/O 비용",
        "BL FEE": "B/L 비용",
        "B/L FEE": "B/L 비용",
        "CLEANING": "클리닝 비용",
        "INSPECTION": "검사 비용",
        "관세": "관세",
        "부가세": "부가세",
        "통관": "통관수수료",
        "운임": "운임",
        "운송": "운송비",
        "창고": "창고료",
        "서류": "서류 발급비",
        "원산지": "원산지 증명서",
    }

    seen_sources: set[str] = set()

    def label_for(line: str) -> str:
        upper = line.upper()
        return next((ko for key, ko in label_map.items() if key in upper), "")

    def classify(line: str, label: str, current_section: str) -> str:
        upper = line.upper()
        if label in {"관세", "부가세"}:
            return "tax"
        if current_section == "tax":
            return "tax"
        if current_section == "logistics":
            return "logistics"
        logistics_words = [
            "FREIGHT",
            "TERMINAL",
            "DOCUMENT",
            "C.F.S",
            "CFS",
            "HANDLING",
            "STORAGE",
            "CERTIFICATE",
            "CUSTOMS",
            "INLAND",
            "TRUCKING",
            "DELIVERY",
            "WAREHOUSE",
            "WHARFAGE",
            "CHARGE",
            "FEE",
            "통관",
            "운임",
            "운송",
            "창고",
            "서류",
            "원산지",
        ]
        if any(word in upper for word in logistics_words):
            return "logistics"
        return ""

    def amount_from(line: str) -> tuple[int, int]:
        amounts = [
            int(raw.replace(",", ""))
            for raw in re.findall(r"(?<![\d.])\d{1,3}(?:,\d{3})+(?![\d.])", line)
        ]
        if not amounts:
            return 0, 0
        amount = amounts[-1]
        vat = 0
        if len(amounts) >= 2 and amounts[-1] < amounts[-2] and "TOTAL" not in line.upper():
            amount = amounts[-2]
            vat = amounts[-1]
        return amount, vat

    def add_cost(group: str, label: str, amount: int, source: str) -> None:
        if group not in {"tax", "logistics"}:
            return
        source_key = f"{group}|{label}|{amount}|{source}"
        if amount <= 0 or source_key in seen_sources:
            return
        seen_sources.add(source_key)
        costs[group].append({"label": label, "amount": amount, "source": source})

    for line in lines:
        if line.startswith("REF NO."):
            parts = line.split()
            if len(parts) >= 3:
                meta["refNo"] = parts[2]
        if line.startswith("H.B/L NO."):
            parts = line.split()
            if len(parts) >= 3:
                meta["hblNo"] = parts[2]
        if "ARRIVAL DATE" in line:
            maybe_date = line.rsplit(" ", 1)[-1]
            if maybe_date.count("-") == 2:
                meta["arrivalDate"] = maybe_date
        if "G.W/T / CBM" in line:
            meta["weightCbm"] = line.replace("G.W/T / CBM", "").strip()

        if "NON INVOICE" in line:
            section = "tax"
            continue
        if "INVOICE :" in line:
            section = "logistics"
            continue
        if line.startswith("SUB TOTAL") or line.startswith("TOTAL") or line.startswith("("):
            continue
        if any(
            skip_word in line.upper()
            for skip_word in [
                "CIF",
                "과세가격",
                "총과세가격",
                "총세액",
                "세액합계",
                "결제금액",
                "제품금액",
                "납부",
                "입금",
                "환율",
                "신고",
                "공급가",
                "합계",
            ]
        ):
            continue

        amount, vat = amount_from(line)
        if not amount:
            continue

        if line.startswith("관세"):
            add_cost("tax", "관세", amount, line)
            continue
        if line.startswith("부가세"):
            add_cost("tax", "부가세", amount, line)
            continue

        label = label_for(line)
        if not label:
            label = re.split(r"\s+KRW\s+|\s+USD\s+|\s+CNY\s+", line, maxsplit=1)[0].strip()
            label = label[:32] or "인보이스 비용"
        group = classify(line, label, section)
        if not group:
            continue
        add_cost(group, label, amount, line)
        if vat:
            add_cost("tax", f"{label} VAT", vat, line)

    totals = {
        "tax": sum(cost["amount"] for cost in costs["tax"]),
        "logistics": sum(cost["amount"] for cost in costs["logistics"]),
        "other": sum(cost["amount"] for cost in costs["other"]),
    }
    totals["grand"] = totals["tax"] + totals["logistics"] + totals["other"]

    return {"meta": meta, "costs": costs, "totals": totals, "rawText": text[:4000]}


def parse_multipart_file(
    headers: Any,
    body: bytes,
    field_name: str = "file",
) -> tuple[bytes, str] | None:
    content_type = headers.get("Content-Type", "")
    if not content_type.startswith("multipart/form-data"):
        return None

    message_bytes = (
        f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8")
        + body
    )
    message = BytesParser(policy=default).parsebytes(message_bytes)
    for part in message.iter_parts():
        disposition = part.get("Content-Disposition", "")
        if "form-data" not in disposition:
            continue
        name = part.get_param("name", header="content-disposition")
        if name != field_name:
            continue
        filename = part.get_filename() or "upload"
        payload = part.get_payload(decode=True) or b""
        return payload, filename
    return None


class Handler(SimpleHTTPRequestHandler):
    def translate_path(self, path: str) -> str:
        if path == "/" or path.startswith("/static/"):
            rel = "index.html" if path == "/" else path.removeprefix("/static/")
            return str(STATIC / rel)
        return str(ROOT / path.lstrip("/"))

    def do_GET(self) -> None:
        if self.path == "/health":
            self._json({"ok": True})
            return
        super().do_GET()

    def do_POST(self) -> None:
        if self.path not in {"/api/import", "/api/import-invoice"}:
            self.send_error(404)
            return

        content_length = int(self.headers.get("Content-Length", "0") or "0")
        if content_length > MAX_UPLOAD_BYTES:
            self.send_error(413, "uploaded file is too large")
            return

        uploaded = parse_multipart_file(
            self.headers,
            self.rfile.read(content_length),
            field_name="file",
        )
        if uploaded is None:
            self.send_error(400, "file field is required")
            return

        try:
            file_bytes, filename = uploaded
            if self.path == "/api/import-invoice":
                payload = parse_invoice_pdf(file_bytes, filename)
            else:
                payload = parse_shipping_list(file_bytes, filename)
        except Exception as exc:  # pragma: no cover - displayed to the local user
            self._json({"error": str(exc)}, status=500)
            return
        self._json(payload)

    def end_headers(self) -> None:
        if self.path.endswith(".js"):
            self.send_header("Content-Type", "text/javascript; charset=utf-8")
        super().end_headers()

    def guess_type(self, path: str) -> str:
        if path.endswith(".js"):
            return "text/javascript"
        return mimetypes.guess_type(path)[0] or "application/octet-stream"

    def _json(self, body: dict[str, Any], status: int = 200) -> None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def main() -> None:
    port = int(os.environ.get("PORT", DEFAULT_PORT))
    host = os.environ.get("HOST", "0.0.0.0")
    server = ReusableThreadingHTTPServer((host, port), Handler)
    print(f"수입원가 계산기: http://{host}:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
